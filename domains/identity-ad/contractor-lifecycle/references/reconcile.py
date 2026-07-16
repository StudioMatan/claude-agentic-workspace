#!/usr/bin/env python3
"""Monthly contractor reconciliation: AD (enabled) vs the vendor rosters.
Usage: python3 reconcile.py "<month folder>"   (folder holds the rosters + the AD pull CSV)
Read-only over sources. Writes Contractor'sMainCheck_<folder>.xlsx into the folder.

Vendor loaders below encode each vendor's real-world spreadsheet quirks (header offsets,
missing email columns, pivot preambles). When a vendor changes layout, the run fails
loudly - update the matching load_*() and note it in the SKILL.md quirks section.
"""
import openpyxl, csv, re, datetime, difflib
from openpyxl.styles import Font, PatternFill, Alignment

import sys, glob, os
BASE=os.path.abspath(sys.argv[1]) if len(sys.argv)>1 else os.getcwd()   # month folder
def _one(pat, what):
    m=sorted(glob.glob(f"{BASE}/{pat}"))
    if not m: sys.exit(f"Cannot find {what} ({pat}) in {BASE}")
    return m[-1]                                       # newest by name/timestamp
AD_CSV=_one("ContractorsActiveusers-*.csv","AD pull CSV")
F_VENDOR_A=_one("Vendor-A*.xlsx","Vendor-A roster")
F_VENDOR_B=_one("Vendor-B*.xlsx","Vendor-B roster")
F_VENDOR_C=_one("Vendor-C*.xlsx","Vendor-C roster")
F_VENDOR_D=_one("Vendor-D*.xlsx","Vendor-D tracker")   # filename spelling drifts - keep glob loose
TODAY=datetime.datetime.now().replace(hour=0,minute=0,second=0,microsecond=0)
TAG=os.path.basename(os.path.abspath(BASE)).replace(" ","")             # e.g. June2026
print(f"Folder: {BASE}")
print(f"AD CSV: {os.path.basename(AD_CSV)}")
MAIN=["Vendor-A","Vendor-B","Vendor-C","Vendor-D"]

def norm_name(s):
    if not s: return ""
    s=str(s).lower()
    s=re.sub(r"\(.*?\)","",s)                # drop parentheticals e.g. "(Returned)"
    s=re.sub(r"[^a-z ]"," ",s)
    toks=[t for t in s.split() if t]
    return " ".join(sorted(toks))            # order-independent name key

def name_tokens(s):
    return set(norm_name(s).split())

def names_match(a,b):
    """token-aware: subset (extra surname) or strong Jaccard, else close fuzzy."""
    if not a or not b: return False
    ta,tb=set(a.split()),set(b.split())
    if not ta or not tb: return False
    if ta<=tb or tb<=ta: return True                    # one is subset of the other
    inter=len(ta&tb); jac=inter/len(ta|tb)
    if jac>=0.6: return True
    return difflib.SequenceMatcher(None,a,b).ratio()>=0.9

def local(email):
    return str(email).strip().lower().split("@")[0] if email and "@" in str(email) else ""

# ---------- AD ----------
ad=[]
with open(AD_CSV, encoding="utf-8-sig") as fh:
    for row in csv.DictReader(fh):
        row={ (k.lstrip("﻿") if k else k):v for k,v in row.items()}
        ad.append(row)
for u in ad:
    u["_name"]=norm_name(u["Name"]); u["_local"]=local(u["Email"])
print(f"AD enabled contractors: {len(ad)}")

# ---------- rosters ----------
def load_vendor_a():
    wb=openpyxl.load_workbook(F_VENDOR_A, read_only=True, data_only=True)
    sname=next((s for s in wb.sheetnames if "active" in s.lower() and "roster" in s.lower()), wb.sheetnames[0]); ws=wb[sname]; rows=list(ws.iter_rows(values_only=True))
    out=[]
    for r in rows[3:]:
        name=r[2]
        if not name or not str(name).strip(): continue
        ended=r[6]  # Date Ended
        if ended: continue                    # ended -> not active
        email=r[15] if len(r)>15 and r[15] and "@" in str(r[15]) else None  # corporate email col
        out.append({"name":name,"email":email,"status":"Active"})
    return out

def load_vendor_b():
    wb=openpyxl.load_workbook(F_VENDOR_B, read_only=True, data_only=True)
    ws=wb["Roster"]; rows=list(ws.iter_rows(values_only=True)); out=[]
    for r in rows[1:]:
        name=r[5]  # Full Name
        if not name or not str(name).strip(): continue
        out.append({"name":name,"email":None,"status":str(r[9] or "")})
    return out

def load_vendor_c():
    """Month-named sheet, person block from the 'Corp Email' header.
    Active = Billing in {Billable, AV, Floater}. 'Deactivated' captured separately."""
    wb=openpyxl.load_workbook(F_VENDOR_C, read_only=True, data_only=True)
    sheet=None
    for s in wb.sheetnames:
        if s.replace(" ","").lower() in TAG.lower() or TAG.lower() in s.replace(" ","").lower():
            sheet=s; break
    if not sheet: sheet=wb.sheetnames[1] if wb.sheetnames[0]=="Current" else wb.sheetnames[0]
    print(f"Vendor-C sheet used: {sheet!r}")
    ws=wb[sheet]; rows=list(ws.iter_rows(values_only=True))
    active=[]; deact=[]; started=False
    for r in rows:
        if not started:
            if r and any(isinstance(c,str) and c.strip()=="Corp Email" for c in r):
                started=True
            continue
        name=r[1] if len(r)>1 else None
        billing=str(r[2]).strip() if len(r)>2 and r[2] else ""
        email=r[3] if len(r)>3 else None
        if not name: continue
        rec={"name":name,"email":email if (email and "@" in str(email)) else None,"status":billing}
        if billing.lower()=="deactivated": deact.append(rec)
        elif billing in ("Billable","AV","Floater"): active.append(rec)
    return active, deact

def load_vendor_d():
    wb=openpyxl.load_workbook(F_VENDOR_D, read_only=True, data_only=True)
    ws=wb["Sheet1"]; rows=list(ws.iter_rows(values_only=True)); out=[]
    for r in rows[1:]:
        fn,ln,email,status=r[0],r[1],r[2],r[3]
        nm=" ".join([str(x).strip() for x in (fn,ln) if x])
        if not nm and not email: continue
        out.append({"name":nm,"email":email,"status":str(status or "")})
    return out

vc_active, vc_deact=load_vendor_c()
rosters={"Vendor-A":load_vendor_a(),"Vendor-B":load_vendor_b(),
         "Vendor-C":vc_active,"Vendor-D":load_vendor_d()}
for v,r in rosters.items():
    for x in r: x["_name"]=norm_name(x["name"]); x["_local"]=local(x["email"])
for x in vc_deact: x["_name"]=norm_name(x["name"]); x["_local"]=local(x["email"])
for v,r in rosters.items(): print(f"Roster {v} (active): {len(r)}")
print(f"Vendor-C marked Deactivated (reference): {len(vc_deact)}")

# ---------- matching ----------
def match_in(target_records, name_key, local_key):
    """email localpart first, then token-aware name match."""
    if local_key:
        for t in target_records:
            if t["_local"] and t["_local"]==local_key: return t
    for t in target_records:
        if names_match(name_key,t["_name"]): return t
    return None

ad_by_vendor={v:[u for u in ad if u["ChildOU"]==v] for v in MAIN}
non_main=[u for u in ad if u["ChildOU"] not in MAIN]

# is this AD user on their vendor's active roster?
def on_active_roster(u):
    v=u["ChildOU"]
    if v not in MAIN: return None            # no roster covers this OU
    return match_in(rosters[v],u["_name"],u["_local"]) is not None

# roster -> AD (active roster entry, no enabled AD account)
roster_not_in_ad=[]
for v in MAIN:
    for x in rosters[v]:
        if v=="Vendor-D" and x["status"].strip().lower()!="active": continue
        if not match_in(ad_by_vendor[v],x["_name"],x["_local"]):
            roster_not_in_ad.append({"vendor":v,**x})

# expiration
def parse_dt(s):
    try: return datetime.datetime.strptime(str(s)[:19],"%Y-%m-%d %H:%M:%S")
    except: return None
for u in ad:
    dt=parse_dt(u["EndDate"]); u["_dt"]=dt
    u["_days"]=(dt-TODAY).days if dt else None
    u["_onroster"]=on_active_roster(u)

# classification
disable=[]     # expired AND (off active roster OR OU has no roster) -> leaver
extend=[]      # expired but still on active roster -> renew expiry
offroster_active=[]  # enabled, NOT expired, but off active roster (main vendors) -> verify
soon=[]
for u in ad:
    d=u["_days"]
    if d is not None and d<0:
        if u["ChildOU"] in MAIN and u["_onroster"]:
            extend.append(u)
        else:
            disable.append(u)                # off-roster or unrostered OU
    elif d is not None and 0<=d<=30:
        soon.append(u)
    elif u["ChildOU"] in MAIN and u["_onroster"] is False:
        offroster_active.append(u)           # enabled, not expiring, off roster
for lst in (disable,extend,soon): lst.sort(key=lambda u:(u["_days"] if u["_days"] is not None else 9999))

print("\n===== RESULTS =====")
print(f"DISABLE (expired & off-roster / unrostered OU): {len(disable)}")
for u in disable: print(f"  {u['ChildOU']:12} | {u['Name']:28} | exp {abs(u['_days'])}d ago | roster={u['_onroster']}")
print(f"\nEXTEND expiry (expired but ON active roster): {len(extend)}")
for u in extend: print(f"  {u['ChildOU']:12} | {u['Name']:28} | exp {abs(u['_days'])}d ago")
print(f"\nOFF-ROSTER but enabled & not expiring (verify): {len(offroster_active)}")
for u in offroster_active: print(f"  {u['ChildOU']:12} | {u['Name']:28} | {u['Email']}")
print(f"\nActive roster entry with NO enabled AD account: {len(roster_not_in_ad)}")
for x in roster_not_in_ad: print(f"  {x['vendor']:10} | {x['name']:28} | {x.get('email')}")
print(f"\nSoon-to-expire (<=30d): {len(soon)}")
for u in soon: print(f"  {u['ChildOU']:12} | {u['Name']:28} | {u['_dt'].date()} | {u['_days']}d | roster={u['_onroster']}")

# ---------- write workbook (3-sheet MainCheck, noted rows first, explicit notes) ----------
wb=openpyxl.Workbook(); wb.remove(wb.active)
YELLOW=PatternFill("solid",fgColor="FFEB9C")   # data update requirement
RED   =PatternFill("solid",fgColor="FFD9D9")   # missing users - review for deactivation
EXPRED=PatternFill("solid",fgColor="FFC7CE")   # expired rows
BOLD=Font(bold=True)

def style_hdr(ws):
    for c in ws[1]: c.font=BOLD
    for col in ws.columns:
        w=max((len(str(c.value)) for c in col if c.value is not None),default=10)
        ws.column_dimensions[col[0].column_letter].width=min(max(w+2,12),70)
    ws.freeze_panes="A2"

def write_sorted(ws,rows):
    """rows = list of (values, fill). Red first, yellow second, clean last."""
    order={id(RED):0,id(EXPRED):0,id(YELLOW):1,id(None):2}
    for vals,fill in sorted(rows,key=lambda t:order.get(id(t[1]),2)):
        ws.append(vals)
        if fill:
            for c in ws[ws.max_row]: c.fill=fill
    style_hdr(ws)

# known cross-spellings: roster spelling -> AD spelling (normalized keys).
# One explicit entry per confirmed case - never fuzzy-guess silently.
ALIAS={"doe jano":"doe jane"}

def find_ad(v,x):
    key=ALIAS.get(x["_name"],x["_name"])
    if x["_local"]:
        for u in ad_by_vendor[v]:
            if u["_local"]==x["_local"]: return u
    for u in ad_by_vendor[v]:
        if names_match(key,u["_name"]): return u
    return None

# --- Sheet 1: Contractors to AD (vendor lists checked against AD)
rows1=[]
for v in MAIN:
    for x in rosters[v]:
        if v=="Vendor-D" and x["status"].strip().lower()!="active": continue
        u=find_ad(v,x)
        note=None; fill=None
        remail=str(x["email"]).strip().lower() if x["email"] else None
        if not u:
            note="Not found in AD - no enabled account. Please confirm this person is still active"; fill=RED
        else:
            ademail=str(u["Email"]).strip() if u["Email"] and "@" in str(u["Email"]) else None
            if ALIAS.get(x["_name"])==u["_name"]:
                note=f"Name is misspelled on your list - our records show: {u['Name']}. Please correct the list"; fill=YELLOW
            elif v!="Vendor-B" and ademail:
                if not remail:
                    note=f"Email missing on your list - please add: {ademail}"; fill=YELLOW
                elif remail!=ademail.lower():
                    if "legacy-brand.com" in remail:   # pre-acquisition domain still on vendor lists
                        note=(f"Your list still shows the old legacy-brand address - the account was migrated. "
                              f"Please update the list to: {ademail}"); fill=YELLOW
                    else:
                        note=f"Email on your list does not match our records - please update to: {ademail}"; fill=YELLOW
        rows1.append(([x["name"],x["email"],v,note],fill))
ws1=wb.create_sheet("Contractors to AD")
ws1.append(["Name","Email on Vendor List","Contractor","Note"])
write_sorted(ws1,rows1)

# --- Sheet 2: AD to Contractors (AD accounts checked against vendor lists)
rows2=[]
for v in MAIN:
    for u in sorted(ad_by_vendor[v],key=lambda z:z["Name"]):
        onr=u["_onroster"]
        if not onr:
            for x in rosters[v]:
                if ALIAS.get(x["_name"])==u["_name"]: onr=True; break
        note=None; fill=None
        if not onr:
            note="Has an active AD account but does NOT appear on your list - please confirm: still active, or should we deactivate?"; fill=RED
        elif str(u["Manager"]).strip().startswith("X"):
            note=f"Listed manager ({u['Manager'][4:]}) has left the company - please provide the current manager"; fill=YELLOW
        rows2.append(([u["Name"],u["Email"],v,u["Manager"],note],fill))
ws2=wb.create_sheet("AD to Contractors")
ws2.append(["Name","Email","Contractor","Manager","Note"])
write_sorted(ws2,rows2)

# --- Sheet 3: Soon-to-Expire (expired first, most overdue on top)
ws3=wb.create_sheet("Soon-to-Expire")
ws3.append(["Name","Email","Contractor","Manager","EndDate","DaysLeft","Status"])
rows3=sorted([u for u in ad if u["_days"] is not None and u["_days"]<=30], key=lambda z:z["_days"])
for u in rows3:
    d=u["_days"]
    if d<0:
        if u["ChildOU"] in MAIN and u["_onroster"]:
            status=f"EXPIRED {-d}d ago - still on vendor list, end date needs extension"
        else:
            status=f"EXPIRED {-d}d ago - not on any vendor list, review for deactivation"
    else:
        status=f"Expires in {d}d"
    ws3.append([u["Name"],u["Email"],u["ChildOU"],u["Manager"],u["_dt"].strftime("%Y-%m-%d"),d,status])
    for c in ws3[ws3.max_row]: c.fill=EXPRED if d<0 else YELLOW
style_hdr(ws3)

OUT=f"{BASE}/Contractor'sMainCheck_{TAG}.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
for ws in (ws1,ws2,ws3): print(f"  {ws.title}: {ws.max_row-1} rows")
